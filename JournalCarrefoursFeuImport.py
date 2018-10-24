import os
from openpyxl import load_workbook
import pandas as pd
import xlrd
import csv

os.chdir("/Users/griceldacalzada/Documents/Python/TestCarrefoursFeux")
DirTestCarrefoursFeu=os.getcwd()



for filename in os.listdir(DirTestCarrefoursFeu+"/Excel"):
    if filename==".DS_Store":
        continue
    else:
        wb=load_workbook("./Excel/"+filename)
        Journal=wb['Journal']
        TempB3S=Journal.cell(row=3,column=2)
        DataB3S=[]
        for i in range (3, Journal.max_row):
            DataLine=[]
            for j in range(1,26):
                DataLine.append(Journal.cell(row=i,column=j).value)
            DataB3S.append(DataLine)
        outfile=open("./CSV/" + filename[:len(filename) - 5] + ".csv", "w")
        writer = csv.writer(outfile, delimiter=';',quotechar='"')
        writer.writerow(DataB3S)
        outfile.close()





















